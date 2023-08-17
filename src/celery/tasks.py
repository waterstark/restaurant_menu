import openpyxl

from src.crud.dish_crud import DishCrud
from src.crud.menu_crud import MenuCrud
from src.crud.submenu_crud import SubMenuCrud
from src.database import async_session_maker
from src.schemas import DishModel, MenuModel, SubmenuModel


def get_data_from_xlsx():
    workbook = openpyxl.load_workbook('src/admin/Menu.xlsx')
    worksheet = workbook.active

    black_list = ['None']
    menu = {}
    submenu = {}
    dish = {}
    for row in range(1, worksheet.max_row + 1):
        dictionary = {}
        if (
            str(worksheet.cell(row, 2).value) not in black_list
            and len(str(worksheet.cell(row, 2).value)) < 35
        ):
            key = 'title'
            value = str(worksheet.cell(row, 2).value)
            dictionary[key] = value
            if (
                str(worksheet.cell(row, 3).value) not in black_list
                and str(worksheet.cell(row, 3).value) != 'None'
            ):
                key = 'description'
                value = str(worksheet.cell(row, 3).value)
                dictionary[key] = value
                key = 'id'
                menu_id = str(worksheet.cell(row, 1).value)
                dictionary[key] = menu_id
            menu[menu_id] = dictionary
        if (
            str(worksheet.cell(row, 3).value) not in black_list
            and len(str(worksheet.cell(row, 3).value)) < 35
            and str(worksheet.cell(row, 4).value) != 'None'
        ):
            key = 'title'
            value = str(worksheet.cell(row, 3).value)
            dictionary[key] = value
            if (
                str(worksheet.cell(row, 4).value) not in black_list
                and str(worksheet.cell(row, 4).value) != 'None'
            ):
                key = 'menu_id'
                dictionary[key] = menu_id
                key = 'description'
                value = str(worksheet.cell(row, 4).value)
                dictionary[key] = value
                key = 'id'
                submenu_id = str(worksheet.cell(row, 2).value)
                dictionary[key] = submenu_id
            submenu[submenu_id] = dictionary
        if (
            str(worksheet.cell(row, 4).value) not in black_list
            and len(str(worksheet.cell(row, 4).value)) < 35
            and str(worksheet.cell(row, 5).value) != 'None'
        ):
            key = 'title'
            value = str(worksheet.cell(row, 4).value)
            dictionary[key] = value
            if (
                str(worksheet.cell(row, 5).value) not in black_list
                and str(worksheet.cell(row, 5).value) != 'None'
            ):
                key = 'submenu_id'
                dictionary[key] = submenu_id
                key = 'description'
                value = str(worksheet.cell(row, 5).value)
                dictionary[key] = value
                key = 'price'
                value = str(worksheet.cell(row, 6).value)
                dictionary[key] = value
                key = 'id'
                value = str(worksheet.cell(row, 3).value)
                dictionary[key] = value
                dish[value] = dictionary
    return menu, submenu, dish


async def get_data_from_db(db_dict):
    menus = {}
    submenus = {}
    dishes = {}
    for menu in db_dict:
        menu_dict = {
            'id': str(menu.id),
            'title': menu.title,
            'description': menu.description,
        }
        menus[str(menu.id)] = menu_dict
        for submenu in menu.submenu:
            submenu_dict = {
                'id': str(submenu.id),
                'title': submenu.title,
                'description': submenu.description,
                'menu_id': str(menu.id),
            }
            submenus[str(submenu.id)] = submenu_dict
            for dish in submenu.dishes:
                dish_dict = {
                    'id': str(dish.id),
                    'title': dish.title,
                    'description': dish.description,
                    'price': dish.price,
                    'submenu_id': dish.submenu_id,
                }
                dishes[str(dish.id)] = dish_dict
    return menus, submenus, dishes


async def comparation():
    async with async_session_maker() as session:
        menu_crud = MenuCrud(session)
        submenu_crud = SubMenuCrud(session)
        dish_crud = DishCrud(session)
        db_data = await menu_crud.get_all_data()

        menus_xlsx, submenus_xlsx, dishes_xlsx = get_data_from_xlsx()
        db_menu, db_submenu, db_dish = await get_data_from_db(db_data)

        for key, values in menus_xlsx.items():
            if db_menu.get(key) is None:
                await menu_crud.create(new_menu=MenuModel(**values), id=key)
            elif db_menu.get(key) != values:
                await menu_crud.update(new_menu=MenuModel(**values), menu_id=key)
        for key, values in db_menu.items():
            if menus_xlsx.get(key) is None:
                await menu_crud.delete(menu_id=key)

        for key, values in submenus_xlsx.items():
            if db_submenu.get(key) is None:
                await submenu_crud.create(
                    new_submenu=SubmenuModel(**values),
                    id=key,
                    menu_id=values['menu_id'],
                )
            elif db_submenu.get(key) != values:
                await submenu_crud.update(
                    new_submenu=SubmenuModel(**values),
                    submenu_id=key,
                    menu_id=values['menu_id'],
                )
        for key, values in db_submenu.items():
            if submenus_xlsx.get(key) is None:
                await submenu_crud.delete(submenu_id=key)

        for key, values in dishes_xlsx.items():
            if db_dish.get(key) is None:
                await dish_crud.create(
                    new_dish=DishModel(**values),
                    id=key,
                    submenu_id=values['submenu_id'],
                )
            elif db_dish.get(key) != values:
                await dish_crud.update(
                    new_dish=DishModel(**values),
                    dish_id=key,
                    submenu_id=values['submenu_id'],
                )
        for key, values in db_dish.items():
            if dishes_xlsx.get(key) is None:
                await dish_crud.delete(dish_id=key)
